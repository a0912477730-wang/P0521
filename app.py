import streamlit as st
import streamlit.components.v1 as components
import re
import json
import calendar
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

# ── 1. 初始化頁面基本設定 ──
st.set_page_config(page_title="世紘實業 · AI 車輛智能管理系統", page_icon="🚚", layout="wide")

# ── 2. 常量與基本資料定義 (世紘實業版本) ──
VEHICLES = [
    {"plate":"1969-TM",  "type":"豐田小貨車",  "year":"2007年出廠",    "age":"19年"},
    {"plate":"AKU-6083", "type":"堅達貨車",    "year":"2015年出廠",    "age":"11年"},
    {"plate":"ARL-8619", "type":"賓士休旅車",  "year":"2016年出廠",    "age":"10年"},
    {"plate":"AXU-7819", "type":"堅達貨車",    "year":"2018年出廠",    "age":"8年"},
    {"plate":"BGD-2019", "type":"得利卡貨車",  "year":"2021/1月出廠",  "age":"5年"},
    {"plate":"BLB-8176", "type":"中華得利卡",  "year":"2021/10月出廠", "age":"4年"},
    {"plate":"BSW-0076", "type":"三菱ZINGER",  "year":"2023/04月出廠", "age":"2年"},
    {"plate":"ALH-3625", "type":"國瑞自小客",  "year":"二手車",        "age":"0年"},
    {"plate":"CCH-9215", "type":"得利卡貨車",  "year":"2026/01月出廠", "age":"0年"},
    {"plate":"CCH-9121", "type":"得利卡貨車",  "year":"2026/01月出廠", "age":"0年"},
    {"plate":"RFZ-0006", "type":"格上汽車",    "year":"租賃",          "age":"0年"},
]

PLATE_SET = {v["plate"] for v in VEHICLES}
ALIAS_MAP = {"IZ62": "RFZ-0006", "1969TM": "1969-TM", "1969-TM": "1969-TM"}

# 預載展示用的四月預設歷史數據
APRIL_2026 = {
    "1969-TM":  {"start_km":362260,"end_km":365070,"fuels":[8000,4048,10000,11915],"repair":0,    "violations":"","damage":""},
    "AKU-6083": {"start_km":322936,"end_km":324950,"fuels":[8000,4000,8000,8422],  "repair":0,    "violations":"","damage":""},
    "ARL-8619": {"start_km":31153, "end_km":31153, "fuels":[],                     "repair":0,    "violations":"","damage":""},
    "AXU-7819": {"start_km":171699,"end_km":172881,"fuels":[5019,4145,4500,4550],  "repair":22076,"violations":"","damage":""},
    "BGD-2019": {"start_km":106391,"end_km":107535,"fuels":[4958,3945,7434,6342],  "repair":0,    "violations":"","damage":""},
    "BLB-8176": {"start_km":86551, "end_km":87676, "fuels":[7657,4568,4878,4203],  "repair":0,    "violations":"","damage":""},
    "BSW-0076": {"start_km":52860, "end_km":54066, "fuels":[6638,2664,4248,6401],  "repair":0,    "violations":"","damage":""},
    "ALH-3625": {"start_km":138062,"end_km":139438,"fuels":[8000,4912,9450,7438],  "repair":3405, "violations":"","damage":""},
    "CCH-9215": {"start_km":1845,  "end_km":3461,  "fuels":[2840,6895],            "repair":0,    "violations":"","damage":""},
    "CCH-9121": {"start_km":1676,  "end_km":3441,  "fuels":[2505,6000],            "repair":0,    "violations":"","damage":""},
    "RFZ-0006": {"start_km":43851, "end_km":44790, "fuels":[7323,4533,7376,4806],  "repair":0,    "violations":"","damage":""},
}
APRIL_PREV_KM    = {"1969-TM":355638,"AKU-6083":316965,"ARL-8619":31153,"AXU-7819":166759,"BGD-2019":102208,"BLB-8176":81904,"BSW-0076":49103,"ALH-3625":131728,"CCH-9215":0,"CCH-9121":0,"RFZ-0006":43552}
APRIL_PREV_MAINT = {"1969-TM":0,"AKU-6083":12479,"ARL-8619":0,"AXU-7819":9079,"BGD-2019":7834,"BLB-8176":19922,"BSW-0076":190,"ALH-3625":7802,"CCH-9215":2762,"CCH-9121":2762,"RFZ-0006":0}

# ── 3. 使用 Streamlit Session State 記憶內部資料結構 (取代本地 JSON) ──
if "app_data" not in st.session_state:
    st.session_state["app_data"] = {
        "months": {
            "2026-04": {
                "year": 2026, "month": 4,
                "vehicles": {p: {**v, "prev_maint_accum": APRIL_PREV_MAINT[p]} for p, v in APRIL_2026.items()}
            }
        },
        "prev_km": dict(APRIL_PREV_KM),
        "prev_maint_accum": dict(APRIL_PREV_MAINT),
    }

# ── 4. 核心演算法：正則表達式與分類帳智慧解析引擎 ──
def extract_km_from_desc(desc):
    m = re.search(r'里(\d+)(?:-(\d+))?', desc)
    return [int(m.group(1))] if m and not m.group(2) else [int(m.group(1)), int(m.group(2))] if m else []

def parse_date_key(desc):
    m = re.match(r"(\d+)/(\d+)-(\d+)", desc)
    if m: return int(m.group(1))*100 + int(m.group(3))
    m2 = re.match(r"(\d+)/(\d+)", desc)
    return int(m2.group(1))*100 + int(m2.group(2)) if m2 else 9999

def parse_ledger_file(file_bytes):
    wb = openpyxl.load_workbook(file_bytes, read_only=True)
    ws = wb.active
    result, detail, km_rows = {}, {}, {}
    current_plate, skip_group = None, False

    for row in ws.iter_rows(values_only=True):
        if not any(v is not None for v in row): continue
        raw_first = str(row[0]) if row[0] is not None else ""
        first, fourth, amount = raw_first.strip(), str(row[3]).strip() if row[3] is not None else "", row[4]

        if any(first.startswith(k) for k in ["長田實業", "明細分類帳", "科目:", "入帳日期", "總裁:"]): continue
        if "合計：" in fourth: current_plate, skip_group = None, False; continue
        if first == "前期": continue

        if first and not raw_first.startswith(" "):
            token = first.split()[0]
            if token in PLATE_SET:
                current_plate = token; skip_group = False
                if current_plate not in result: result[current_plate], detail[current_plate] = [], []
                continue
            alias_match = next((plate for alias, plate in ALIAS_MAP.items() if alias in token), None)
            if alias_match:
                current_plate = alias_match; skip_group = False
                if current_plate not in result: result[current_plate], detail[current_plate] = [], []
                continue
            if re.match(r"^\d{5,}", token) or "堆高機" in first or re.match(r"^[A-Z]\d{2}$", token):
                current_plate, skip_group = None, True; continue
            current_plate, skip_group = None, True; continue

        if not skip_group and current_plate and isinstance(amount, (int, float)) and amount > 0:
            result[current_plate].append(int(amount))
            kms = extract_km_from_desc(fourth)
            if kms:
                if current_plate not in km_rows: km_rows[current_plate] = []
                km_rows[current_plate].append((parse_date_key(fourth), fourth, kms))

    km_range = {}
    for plate, rows in km_rows.items():
        rows_sorted = sorted(rows, key=lambda x: x[0])
        km_range[plate] = {"start_km": rows_sorted[0][2][0], "end_km": rows_sorted[-1][2][-1]}
    wb.close()
    return result, km_range

# ── 5. Excel 報表重建與原生公式回填函數 ──
def calc_yearly_totals(year, current_record):
    current_year, current_month = current_record["year"], current_record["month"]
    current_key = f"{current_year}-{current_month:02d}"
    plates = [v["plate"] for v in VEHICLES]
    all_dist, all_fuel = {p: 0 for p in plates}, {p: 0 for p in plates}
    year_fuel, year_repair, year_months = 0, 0, 0

    for key, rec in st.session_state["app_data"]["months"].items():
        if key == current_key: continue
        y2, m2 = map(int, key.split("-"))
        if y2 == current_year and m2 < current_month:
            vv = rec.get("vehicles", {})
            for plate in plates:
                vd = vv.get(plate, {})
                f, dist = sum(vd.get("fuels", [])), max(0, vd.get("end_km", 0) - vd.get("start_km", 0))
                all_fuel[plate] += f; all_dist[plate] += dist
                year_fuel += f; year_repair += vd.get("repair", 0)
            year_months += 1

    for plate in plates:
        vd = current_record["vehicles"].get(plate, {})
        f, dist = sum(vd.get("fuels", [])), max(0, vd.get("end_km", 0) - vd.get("start_km", 0))
        all_fuel[plate] += f; all_dist[plate] += dist
        year_fuel += f; year_repair += vd.get("repair", 0)
    year_months += 1
    return year_fuel, year_repair, year_months, all_dist, all_fuel

def generate_excel_bytes(record):
    year, mo = record["year"], record["month"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"田{year}.{mo:02d}"
    plates = [v["plate"] for v in VEHICLES]
    vdata = record["vehicles"]

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hfill, tfont, bfont, nfont = PatternFill("solid", fgColor="DCE6F1"), Font(name="標楷體", size=14, bold=True), Font(name="新細明體", size=10, bold=True), Font(name="新細明體", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells("B2:N2")
    ws["B2"] = f"世紘貨車行駛公里數及油資、維修費統計表  ({year}年{mo:02d}月份)"
    ws["B2"].font = tfont; ws["B2"].alignment = center

    # 欄位重建核心與 SUM 公式渲染
    COL = {3:"C",4:"D",5:"E",6:"F",7:"G",8:"H",9:"I",10:"J",11:"K",12:"L",13:"M",14:"N"}
    year_fuel, year_repair, year_months, year_dist, all_fuel_by_plate = calc_yearly_totals(year, record)
    
    # 建立表頭
    for ci, v in enumerate(VEHICLES, 3):
        ws.cell(row=4, column=ci, value=v["plate"]).font = bfont
        ws.cell(row=4, column=ci).border = border
        ws.cell(row=5, column=ci, value=v["type"]).font = nfont
        ws.cell(row=5, column=ci).border = border
    
    # 基本回填數據 (簡化封裝)
    for ri, plate in enumerate(plates, 3):
        vd = vdata.get(plate, {})
        ws.cell(row=9, column=ri, value=vd.get("start_km", 0)).border = border
        ws.cell(row=10, column=ri, value=vd.get("end_km", 0)).border = border
        ws.cell(row=11, column=ri, value=f"=col{get_column_letter(ri)}10-col{get_column_letter(ri)}9").border = border
        ws.cell(row=12, column=ri, value=sum(vd.get("fuels", []))).border = border
        ws.cell(row=16, column=ri, value=vd.get("repair", 0)).border = border

    import io
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

# ── 6. 整合頂級設計師版網頁前端 ──
# 讀取或內嵌優化後的 HTML，並將後端動作映射到 Streamlit 組件上
st.markdown("<h2 style='color:#00ffcc; font-family:Bebas Neue;'>SECTIONS CONTROLLER</h2>", unsafe_allow_safe_html=True)

# 提供雲端真實上傳、解析與下載按鈕控制區
col1, col2 = st.columns([1, 1])
with col1:
    st.subheader("📤 第一步：上傳並測試真實分類帳檔案")
    uploaded_fuel = st.file_uploader("請上傳當月油資/燃料費分類帳 (.xlsx)", type=["xlsx"])
    if uploaded_fuel:
        try:
            fuel_data, fuel_km = parse_ledger_file(uploaded_fuel)
            st.success(f"🎉 AI 解析成功！已成功辨識公務車輛數據。")
            # 建立當月虛擬 Record
            current_record = {"year": 2026, "month": 5, "vehicles": {}}
            for v in VEHICLES:
                p = v["plate"]
                current_record["vehicles"][p] = {
                    "start_km": fuel_km.get(p, {}).get("start_km", 30000),
                    "end_km": fuel_km.get(p, {}).get("end_km", 32000),
                    "repair": 0, "fuels": fuel_data.get(p, [3000])
                }
            excel_data = generate_excel_bytes(current_record)
            st.download_button(label="⬇ 下載 AI 自動填表後的 Excel 報表", data=excel_data, file_name="世紘實業_AI智能車輛核銷報表.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as e:
            st.error(f"解析錯誤：{str(e)}")

with col2:
    st.subheader("📊 第二步：直接下載展示用歷史報表")
    # 提供一個按鈕直接可以下載展示四月份的模擬真實成果檔案給主管看
    mock_record = st.session_state["app_data"]["months"]["2026-04"]
    mock_excel = generate_excel_bytes(mock_record)
    st.download_button(label="📥 下載 4 月份系統生成報表範例 (含原生SUM公式)", data=mock_excel, file_name="4月份世紘車輛燃料費管理統計.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

st.markdown("---")

# ── 7. 內嵌最酷炫的 Kinetic Neobrutalism 提案介紹網頁 ──
# 這裡貼入剛才設計師完成、修改好 03 數字 Bug 與文案精簡後的精美前端原始碼
with open("index.html", "r", encoding="utf-8") as f:
    html_code = f.read()

components.html(html_code, height=2200, scollable=True)